import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- Config ---
folder_path = '../artifacts/reference_datasets'
DATASET_PARTITION = 70
random_seed = 42
output_dir = 'generated_datasets'
os.makedirs(output_dir, exist_ok=True)

# --- Load and combine Excel files ---
excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]
dfs = {file: pd.read_excel(os.path.join(folder_path, file)) for file in excel_files}
combined_df = pd.concat(dfs.values(), ignore_index=True)

# --- Separate and shuffle ---
df_true = combined_df[combined_df['is_requirement'] == True].sample(frac=1, random_state=random_seed).reset_index(drop=True)
df_false = combined_df[combined_df['is_requirement'] == False].sample(frac=1, random_state=random_seed).reset_index(drop=True)

min_dataset_amount = min(len(df_true), len(df_false))
generated_dataset_num = min_dataset_amount // DATASET_PARTITION

# --- Violation columns
quality_columns = ['ambiguity_violation', 'feasibility_violation', 'singularity_violation', 'verifiability_violation']

# --- Add dropdowns using working method
def add_dropdowns(filepath, col_names):
    wb = load_workbook(filepath)
    ws = wb.active

    # Get header row mapping
    headers = {cell.value.strip(): idx + 1 for idx, cell in enumerate(ws[1])}

    for col_name in col_names:

        col_letter = get_column_letter(headers[col_name])

        dropdown_options = ["TRUE", "FALSE"]
        dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_options)}"')
        ws.add_data_validation(dv)

        for row in range(2, ws.max_row + 1):
            dv.add(f"{col_letter}{row}")

    for col_name, idx in headers.items():
        col_letter = get_column_letter(idx)

        if col_name == "Sentence":
            ws.column_dimensions[col_letter].width = 50
            wrap = True
        else:
            ws.column_dimensions[col_letter].width = 25
            wrap = False

        for row in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row}"].alignment = Alignment(wrap_text=wrap)

    wb.save(filepath)
    print(f"Generated File: {filepath}")

# --- Generate datasets ---
for i in range(generated_dataset_num):

    half = DATASET_PARTITION // 2
    true_is_requirement_chunk = df_true.iloc[i * half: (i + 1) * half]
    false_is_requirement_chunk = df_false.iloc[i * half: (i + 1) * half]

    quality_chunk = df_true.iloc[i * DATASET_PARTITION : (i + 1) * DATASET_PARTITION]

    sample_is_requirement_combined = pd.concat([true_is_requirement_chunk, false_is_requirement_chunk], ignore_index=True).sample(frac=1, random_state=random_seed + i)

    datasets = {
        'unlabeled_req': sample_is_requirement_combined[['Sentence']].assign(is_requirement=''),
        'labeled_req': sample_is_requirement_combined[['Sentence', 'is_requirement']],
        'quality_labeled': quality_chunk[['Sentence'] + quality_columns],
        'quality_unlabeled': quality_chunk[['Sentence']].assign(**{col: '' for col in quality_columns})
    }

    for name, df in datasets.items():
        file_path = '../artifacts/generated_datasets/' + f'dataset_{i+1}_{name}.xlsx'
        df.to_excel(file_path, index=False)

        if 'req' in name:
            add_dropdowns(file_path, ['is_requirement'])
        if 'quality' in name:
            add_dropdowns(file_path, quality_columns)
