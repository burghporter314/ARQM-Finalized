from app.util.model_impl.wrappers.bert_wrapper import BertWrapper

from openpyxl.utils import get_column_letter
import pandas as pd

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill


class GenerativeQualityModel:

    def __init__(self, content=None, sentences=None):
        """
        Initializes a GenerativeQualityModel object
        """

        self.bert_wrapper = BertWrapper(violation_threshold=.70, display_ngram_summary=False)

    def getOverallQuality(self, texts, batch_size=32, visualize_pdf = True):
        return self.bert_wrapper.predict_requirement_v2(texts, visualize_pdf=visualize_pdf)

    def getExcelResult(self, requirement_flags, requirements_to_analyze):

        result = self.getOverallQuality(requirements_to_analyze, visualize_pdf=False)
        aggregated = {
            i: {
                'Sentence': sentence,
                'is_requirement': requirement_flags[i] == 1,
                'ambiguity_violation': None,
                'feasibility_violation': None,
                'singularity_violation': None,
                'verifiability_violation': None
            }
            for i, sentence in enumerate(requirements_to_analyze)
        }

        # Step 4: Populate quality flags using violation probability logic
        for i, probs in enumerate(result):

            aggregated[i]['ambiguity_violation'] = probs[0] >= .7
            aggregated[i]['feasibility_violation'] = probs[1] < .3
            aggregated[i]['singularity_violation'] = probs[2] < .3
            aggregated[i]['verifiability_violation'] = probs[3] < .3

        df = pd.DataFrame(aggregated.values())

        excel_path = "requirement_quality_analysis.xlsx"
        df.to_excel(excel_path, index=False)

        wb = load_workbook(excel_path)
        ws = wb.active

        # Step 3: Define the table range
        num_cols = len(df.columns)
        num_rows = len(df) + 1  # +1 for header
        end_col_letter = chr(ord('A') + num_cols - 1)
        table_range = f"A1:{end_col_letter}{num_rows}"

        # Step 4: Create and style the Excel table
        table = Table(displayName="RequirementQualityTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        # Step 5: Add table and save
        ws.add_table(table)

        sentence_col_idx = df.columns.get_loc("Sentence") + 1  # 1-based index
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=sentence_col_idx, max_col=sentence_col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        col_letter = get_column_letter(sentence_col_idx)
        ws.column_dimensions[col_letter].width = 80

        for col_idx in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        wb.save(excel_path)